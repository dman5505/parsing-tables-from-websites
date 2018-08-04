import csv
import urllib.request
import xlrd
from bs4 import BeautifulSoup

import openpyxl

f = open('dataoutput.csv', 'w', newline = '')
writer = csv.writer(f)
workbook = xlrd.open_workbook("county.xls")
worksheet = workbook.sheet_by_name("COUNTY")
low = 0
while low < 100:
    url = format(worksheet.cell(low,0).value)
    soup = BeautifulSoup(urllib.request.urlopen("http://sasweb.unc.edu/cgi-bin/broker?_service=default&_program=cwweb.tbFstMon.sas&county=" + url + "&label=County&fn=Ctgr1&format=html&entry=14").read(), 'lxml')
    tbody = soup('table', {"class":"table"})[0].find_all('tr')
    for row in tbody:
        cols = row.findChildren(recursive=False)
        cols = [ele.text.strip() for ele in cols]
        cols.extend([url])
        writer.writerow(cols)
        print(cols)
    low = low + 1

f.close()

low = 0
while low < 100:
    url = format(worksheet.cell(low,0).value)

    wb = openpyxl.Workbook()
    ws = wb.active

    file = open('dataoutput.csv')
    reader = csv.reader(file, delimiter=',')

    for row in reader:
        if row[-1] == url:
            row = row[:-1]
            ws.append(row)
        else:
            continue

    file.close()
    wb.save(url + '_#OfChildreninCustody' + '.xls')
    low = low + 1
